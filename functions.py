import pandas as pd
import requests
import time
import datetime as dt
import random
import os
import copy
# For logging
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, PatternFill, Alignment
import csv
from bs4 import BeautifulSoup  # for realtime price
import finnhub  # for news sentiment

now = dt.datetime.now  # format: 2021-07-15 12:02:53.453282


def AlphaVantage_keygen():  # scrape for more keys?
    keys = ['H11ML6UV21JYK8KF', '3TFR9FM3AP7I10LT']
    # returns a random key between the index range (create process to filter the used tickers)
    return keys[random.randint(0, len(keys)-1)]  # replace 0 with len - len


# Classes - Orderflow, Crypto, Stocks
class OrderFlow():  # Backtest OrderFlow
    def __init__(self):
        self.filled = False
        self.quantity = 0
        self.highest = 0
        self.log_data = []

    def updateHighest(self, price):
        if price > self.highest:
            self.highest = price

    def buy(self, price, time, balance):
        self.filled = True
        self.quantity = balance/price
        # for logging onto xl file
        self.log_data.append(time)
        self.log_data.append(price)
        self.log_data.append(balance)
        #print('BUY@: {} FOR PRICE {} AND CURRENT BALANCE {}'.format(time, price, balance))
        return

    def sell(self, price, time, balance):
        balance = self.quantity * price

        # for logging
        self.log_data.append(time)
        self.log_data.append(price)
        self.log_data.append(balance)
        # for returning
        log_data = copy.deepcopy(self.log_data)
        # reset
        self.filled = False
        self.quantity = 0
        self.highest = 0
        self.log_data = []

        #print('SELL@: {} FOR PRICE {} AND CURRENT BALANCE {}'.format(time, price, balance))
        return balance, log_data


# Crypto
class Crypto:
    def __init__(self, coin, currency, balance):
        self.currency, self.coin, self.balance = currency.upper(), coin.upper(), balance

    def df_historical(self, outputsize, interval):
        # Send request and convert response to JSON
        url = 'https://www.alphavantage.co/query?function=CRYPTO_INTRADAY&symbol={}&market={}&interval={}min&outputsize={}&apikey={}'.format(
            self.coin, self.currency, interval, outputsize, AlphaVantage_keygen())
        response = requests.get(url)
        intraday = response.json()  # gets intra day price data with 1 min period

        # DF property shift for functionality
        data = intraday
        df = pd.DataFrame.from_dict(
            data['Time Series Crypto ({}min)'.format(interval)], orient='index')  # parse the time series dict to the dataframe
        df = df.loc[::-1]  # invert the dataframe
        df.reset_index(inplace=True)  # convert index to a column
        df = df.rename(columns={'index': 'time'})  # rename the index to time
        df['4. close'] = df['4. close'].astype(
            float)  # convert close prices to a float
        return df

    def df_current(self, df):  # update df with current exchange rate
        url = 'https://www.alphavantage.co/query?function=CURRENCY_EXCHANGE_RATE&from_currency={}&to_currency={}&apikey={}'.format(
            self.coin, self.currency, AlphaVantage_keygen())
        r = requests.get(url)
        data = r.json()  # convert the response to JSON

        # convert it to a a dictionary that can be fed to the dataframe
        data, price, time, bid, ask = data['Realtime Currency Exchange Rate'], data[
            '5. Exchange Rate'], data['6. Last Refreshed'], data['8. Bid Price'], data['9. Ask Price']

        # create new dictionary with same format as historical df data
        exchange = {'time': time, '1. open': ask, '2. high': bid,
                    '3. low': None, '4. close': price, '5. volume': None}

        df = df.append(exchange, ignore_index=True)  # add it to the df
        df['4. close'] = df['4. close'].astype(float)
        return df


# Stocks
class Stock:
    def __init__(self, ticker, balance):
        self.ticker, self.balance = ticker.upper(), balance

    def df_historical_intraday(self, interval, outputsize):
        ticker = self.ticker
        url = "https://www.alphavantage.co/query?function=TIME_SERIES_INTRADAY&symbol={}&interval={}min&outputsize={}&apikey={}".format(
            ticker.upper(), interval, outputsize, AlphaVantage_keygen())
        response = requests.get(url)  # add error handling code
        intraday = response.json()

        # adding to DF
        df = pd.DataFrame.from_dict(intraday['Time Series ({}min)'.format(
            str(interval))], orient='index')  # parse the time series dict to the dataframe
        df = df.loc[::-1]  # invert the dataframe
        df.reset_index(inplace=True)  # convert index to a column
        df = df.rename(columns={'index': 'time'})  # rename the index to time
        df['4. close'] = df['4. close'].astype(
            float)  # convert close prices to a float
        return df

    def df_historical_intraday_extended(self, interval):
        ticker = self.ticker.upper()
        url = "https://www.alphavantage.co/query?function=TIME_SERIES_INTRADAY_EXTENDED&symbol={}&interval={}min&slice=year1month12&apikey={}".format(
            ticker.upper(), interval, AlphaVantage_keygen())
        df = pd.read_csv(url)
        return df

    def df_historical_daily(self, outputsize):
        ticker = self.ticker
        url = "https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol={}&outputsize={}&apikey={}".format(
            ticker.upper(), outputsize, AlphaVantage_keygen())
        # add error handling code (with the response)
        response = requests.get(url)
        intraday = response.json()
        # adding to DF
        # parse the time series dict to the dataframe
        df = pd.DataFrame.from_dict(
            intraday['Time Series (Daily)'], orient='index')
        df = df.loc[::-1]  # invert the dataframe
        df.reset_index(inplace=True)  # convert index to a column
        df = df.rename(columns={'index': 'time'})  # rename the index to time
        df['4. close'] = df['4. close'].astype(
            float)  # convert close prices to a float
        return df

    def df_current(self, df, price):
        # google finance scraping
        header = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.164 Safari/537.36"}
        price = None
        ticker = self.ticker
        ticker = ticker.upper()

        # get exchange
        url = "https://finance.yahoo.com/quote/{}?p={}&.tsrc=fin-srch".format(
            ticker, ticker)
        response = requests.get(url).text
        soup = BeautifulSoup(response, 'lxml')
        exchange = soup.find(
            'div', class_="C($tertiaryColor) Fz(12px)").text.split()[0]

        # get price
        url = "https://www.google.com/finance/quote/{}:{}".format(
            ticker.upper(), exchange)
        response = requests.get(url).text  # reponse works (200)
        soup = BeautifulSoup(response, "lxml")
        data = soup.find('div', class_="YMlKec fxKbKc")
        if data != None:
            price = data.text[1:]

        # set up data for df
        exchange = {'time': now(), '1. open': None, '2. high': None,
                    '3. low': None, '4. close': price, '5. volume': None}
        df = df.append(exchange, ignore_index=True)  # add it to the df
        df['4. close'] = df['4. close'].astype(
            float)  # convert prices to floats
        return df

    def df_current_yahoo(self, df):
        ticker = self.ticker
        url = "https://finance.yahoo.com/quote/{}?p={}&.tsrc=fin-srch".format(
            ticker, ticker)
        response = requests.get(url).text
        soup = BeautifulSoup(response, 'lxml')
        price = soup.find(
            'span', class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").text
        appendData = {'time': now(), '4. close': price}
        df = df.append(appendData, ignore_index=True)  # add it to the df
        df['4. close'] = df['4. close'].astype(
            float)  # convert prices to floats
        return df

# Stock Screeners


def MostVolatile():
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=111&s=ta_mostvolatile"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def MostActive():
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=111&s=ta_mostactive"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def TopGainers():
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=111&s=ta_topgainers&ft=4"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def SideChannel():
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=110&s=ta_p_channel"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def SidewaysHighVolume():  # based on volume
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=110&s=ta_p_channel&o=-volume"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def recentNews(ticker):
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/quote.ashx?t={}".format(ticker.upper())
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    return df


def MidRiskScreenerVOL():
    headers = {'User-Agent': 'Mozilla/5.0'}
    url = "https://finviz.com/screener.ashx?v=111&f=ta_pattern_horizontal2,ta_volatility_mo5&ft=3&o=-volume"
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


def horizontalNearSupRes():
    headers = {'User-Agent': 'Mozilla/5.0'}
    #url = "https://finviz.com/screener.ashx?v=110&s=ta_p_horizontal&ar=180"
    url = "https://finviz.com/screener.ashx?v=110&s=ta_p_horizontal&o=-&ar=180"  # decending
    response = requests.get(url, headers=headers).text
    df = pd.read_html(response)[-2]
    df.columns = df.iloc[0]
    df = df[1:]
    df.drop("No.", axis=1, inplace=True)
    return df


# Fundemtnal, News and Analyst
def companyOverview(ticker):
    url = "https://www.alphavantage.co/query?function=OVERVIEW&symbol={}&apikey={}".format(
        ticker.upper(), AlphaVantage_keygen())
    response = requests.get(url).json()
    return response


def newsSentiment(ticker):
    finnhub_client = finnhub.Client(api_key="c3163siad3idae6u9eb0")
    sentiment = finnhub_client.news_sentiment(ticker)
    return sentiment

# Technical Indicator Functions - Adds the indicators to the pandas df containing price data
def SMA(df, period):  # Simple Moving Average
    df['{}MA'.format(period)] = df['{}'.format(
        '4. close')].rolling(period).mean()
    return


def RSI(df, period):  # Relative Strength Index
    # converts all close prices str -> float
    df['4. close'] = df['4. close'].astype(float)
    delta = df['4. close'].diff()  # diff of the current - previous
    up = delta.clip(lower=0)  # if its lower than 0, clip to 0
    down = -1*delta.clip(upper=0)  # neg to pos, upper than 0 clip
    rsi_period = period-1  # period= period + 1 cause index from 0
    ema_up = up.ewm(com=rsi_period, adjust=False).mean()
    ema_down = down.ewm(com=rsi_period, adjust=False).mean()
    rs = ema_up/ema_down
    df['RSI'] = 100 - (100/(1+rs))  # gets added as a float
    return


def EMA(df, period):  # Expotential Moving Average
    df['{}EMA'.format(period)] = df['4. close'].ewm(
        span=period, adjust=False).mean()
    return


def MACD(df):  # Moving Average Convergence Divergence
    # 12 EMA - 26 EMA
    ema12 = df['4. close'].ewm(span=12, adjust=False).mean()
    ema26 = df['4. close'].ewm(span=26, adjust=False).mean()
    macd_calc = ema12 - ema26
    df['MACD'] = macd_calc
    signal = macd_calc.ewm(span=9, adjust=False).mean()
    df['MACDEMA'] = signal
    return


def BOLBANDS(df, period):
    df['ma_20'] = df['4. close'].rolling(period).mean()  # moving average
    df['std'] = df['4. close'].rolling(period).std()  # standard deviation
    df['upper_bolband'] = df['ma_20'] + (2*df['std'])
    df['lower_bolband'] = df['ma_20'] - (2*df['std'])
    return


# Technical Analysis - Realtime
def TA_MACD(df):  # add MACD Correction function, Optional variable
    # True for buy, False for sell, None for hold
    signal = None  # for error handling
    # Get variables from the last row
    last_macd = df.iloc[-1]['MACD']
    last_macd_signal = df.iloc[-1]['MACDEMA']

    # Analysis Algorithm
    if last_macd > last_macd_signal:
        signal = True
    elif last_macd < last_macd_signal:
        signal = False
    return signal


def TA_RSI(df, oversold, overbought):
    signal = None  # Hold
    # Retreive Dataframe Data
    last_rsi = df.iloc[-1]['RSI']
    # Algorithm
    if last_rsi >= overbought:
        signal = False  # Sell
    elif last_rsi <= oversold:
        signal = True  # Buy
    return signal


def TA_DATA_CROSS(df, upper_indicator, lower_indicator, upper, lower):
    signal = None
    # upper and lower are EMA periods (as an int)
    # indicator would be something like EMA,SMA ETC.

    # buys when upper crosses above lower
    if df['{}{}'.format(upper_indicator, str(upper))] > df['{}{}'.format(lower_indicator, str(lower))]:
        signal = True
    else:  # sells when upper below lower
        signal = False
    return signal


def TA_TRAILLING(highest, price, percent):
    # when the price starts going down, if it goes down x% sell
    signal = False  # True to sell, False to hold
    sell_price = highest - (highest * percent)
    if price <= sell_price:
        signal = True
    return signal


def TA_SINGLE(df, indicator):
    signal = None
    price = df.iloc[-1]['4. close']
    indicatorPrice = df.iloc[-1][indicator]
    if price > indicatorPrice:
        signal = True  # buy
    else:
        signal = False  # sell
    return signal


# Technical Analysis - Backtesting
def TA_B_MACD(macd, macd_ema):
    macd_signal = None
    if macd > macd_ema:
        macd_signal = True  # Buy
    elif macd < macd_ema:
        macd_signal = False  # Sell
    return macd_signal


def TA_B_RSI(rsi, overbought, oversold):
    rsi_signal = None
    if rsi >= overbought:
        rsi_signal = False  # overbought/sell
    elif rsi <= oversold:
        rsi_signal = True  # oversold/buy
    return rsi_signal


def TA_B_EMA(close, ema):
    ema_signal = None
    if close > ema:
        ema_signal = True  # buy
    else:
        ema_signal = False  # sell
    return ema_signal


def TA_B_BB(close, upperband, lowerband):  # bol bands
    signal = None
    if close >= upperband:
        signal = False  # sell
    elif close <= lowerband:
        signal = True
    return signal


# Logging functions
def log_new(filename):  # needs now, add var ticker or name as an optional file name
    # To be called before the loop
    if filename is None:
        filename = str(now())[0:10] + '.xlsx'  # overwrites if exists
    wb = Workbook()  # init a workbook
    ws = wb.active  # gets default sheet 1
    ws.append(['Buy(Time)', 'Buy(Price)', 'Sell(Time)',
               'Sell(Price)', 'Diff. Bal.'])
    for col in range(1, 6):
        ws[get_column_letter(col)+'1'].font, ws[get_column_letter(col) +
                                                '1'].fill, ws[get_column_letter(col)+'1'].alignment = Font(bold=True), PatternFill("solid", fgColor="00FFCC00"), Alignment(horizontal="center", vertical="center")
    wb.save(filename)  # creates the file

    # Code for setting up CSV file
    # create new csv file, if exists replace with new shit
    csv_file_name = '{}.csv'.format(str(now())[0:10])
    csv_file = open(csv_file_name, 'w', newline="")
    col_head = ('Buy(Date)', 'Buy(Price)', 'Sell(Date)',
                'Sell(Price)', 'Diff. Bal.')
    writer = csv.writer(csv_file)
    writer.writerow(col_head)
    csv_file.close()

    return filename, csv_file_name


def log_sell(filename, log_data, csv_file_name):
    wb = load_workbook(filename)
    ws = wb.active
    # the last two entries of the log_data is the init diff and final
    delta_balance = log_data[-1]-log_data[-4]
    # Append the data list to the worksheet
    log_data.pop(-1)
    log_data.pop(-3)
    log_data.append(delta_balance)
    ws.append(log_data)
    if delta_balance < 0:  # to color loss rows
        ws['E'+str(ws.max_row)].fill = PatternFill("solid", fgColor="00FF0000")
    wb.save(filename)

    # for CSV file
    csv_file = open(csv_file_name, 'a', newline="")
    data = (log_data[0], log_data[1], log_data[2], log_data[3], log_data[4])
    writer = csv.writer(csv_file)
    writer.writerow(data)
    csv_file.close()
    return


def log_close(filename, balance, fin_balance):
    wb = load_workbook(filename)
    ws = wb.active
    delta_balance = fin_balance - balance  # profit/loss
    ws['H3'], ws['I3'], ws['H4'], ws['I4'], ws['H5'], ws['I5'] = "Init. Bal. :", balance, "Fin. Bal. :", fin_balance, "Net Ch. :", delta_balance
    ws['H3'].fill, ws['H4'].fill, ws['H5'].fill, ws['H3'].font, ws['H4'].font, ws['H5'].font = PatternFill("solid", fgColor="00FFCC00"), PatternFill(
        "solid", fgColor="00FFCC00"), PatternFill("solid", fgColor="00FFCC00"), Font(bold=True), Font(bold=True), Font(bold=True)
    if delta_balance < 0:
        ws['I5'].fill = PatternFill("solid", fgColor="00FF0000")
    else:
        ws['I5'].fill = PatternFill("solid", fgColor="0000FF00")
    wb.save(filename)
    return


# CSV - Scanning Log
def NewScannerLogger(filename):
    ScannerFile = open(filename, 'w', newline="")
    col_head = ('Ticker', 'Algorithmic Net Change', 'Algorithmic % Change',
                'Non Algorithmic Net Change', 'Non Algorithmic % Change', 'Net % Diff')
    writer = csv.writer(ScannerFile)
    writer.writerow(col_head)
    ScannerFile.close()
    return


def AppendScannerLogger(ScannerFileName, data):
    csv_file = open(ScannerFileName, 'a', newline="")
    tupData = (data[0], data[1], data[2], data[3], data[4], data[5])
    writer = csv.writer(csv_file)
    writer.writerow(tupData)
    csv_file.close()
    return
